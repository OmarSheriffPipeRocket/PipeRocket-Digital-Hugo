"""
Build the full audit as a multi-sheet Excel workbook (mirrors the PDF sections).

Reads audit/content_map_audit.json + audit/_judgment_deep.json + GSC + Semrush
(reused from the PDF/CSV builders). Writes audit/PipeRocket-Content-Map-SEO-Audit.xlsx
with one categorized tab per section. Pure data (no formulas).

Usage: python3 scripts/audit_workbook.py
"""

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from audit_content_map import load_redirects, parse_content_map
from audit_master_csv import load_judgment
from audit_pdf import gsc_aggregate, SEMRUSH_AS, SEMRUSH_RD, SEMRUSH_RD_DIST

ROOT = Path(__file__).resolve().parent.parent
AUDIT = ROOT / "audit"
OUT = AUDIT / "PipeRocket-Content-Map-SEO-Audit.xlsx"
ARTICLE = {"blog", "glossary", "list", "compare", "alternative", "case-study"}

NAVY = "0B2440"
ACCENT = "0BA6E2"
LIGHT = "F1F8FC"
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
CELL_FONT = Font(name="Arial", size=9)
TITLE_FONT = Font(name="Arial", bold=True, color="0B2440", size=13)
SUB_FONT = Font(name="Arial", bold=True, color="0BA6E2", size=11)


def sheet(wb, name, headers, rows, widths=None, wrap=()):
    ws = wb.create_sheet(name[:31])
    ws.append(headers)
    for c in ws[1]:
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.font = HEADER_FONT
        c.alignment = Alignment(vertical="center", wrap_text=True)
    for r in rows:
        ws.append(["" if v is None else v for v in r])
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.font = CELL_FONT
            c.alignment = Alignment(vertical="top", wrap_text=(c.column_letter in wrap))
    ws.freeze_panes = "A2"
    if rows:
        ws.auto_filter.ref = ws.dimensions
    for col, w in (widths or {}).items():
        ws.column_dimensions[col].width = w
    return ws


def main():
    pages = json.loads((AUDIT / "content_map_audit.json").read_text())["pages"]
    rh = json.loads((AUDIT / "content_map_audit.json").read_text()).get("redirect_hygiene", {})
    cmeta = parse_content_map()
    anchors, keywords, cannib = load_judgment()
    g = gsc_aggregate()
    kw = [p for p in pages if p["keyword_target"]]
    by = {p["url"]: p for p in pages}

    wb = Workbook()
    wb.remove(wb.active)

    # ---- 0. Contents ----
    ws = wb.create_sheet("Contents")
    ws["A1"] = "PipeRocket — Content-Map SEO Audit"; ws["A1"].font = TITLE_FONT
    ws["A2"] = "Companion workbook to the PDF. One tab per section. Data snapshot: qp_rollup_2026-06-08."
    ws["A2"].font = CELL_FONT
    ws["A4"] = "HEADLINE: the site doesn't rank for non-brand terms because of an AUTHORITY deficit, not on-page issues."
    ws["A4"].font = Font(name="Arial", bold=True, color="C0392B", size=10)
    ws["A5"] = ("Authority Score 20; 59% of referring domains are AS 0-10, only ~19 are AS 41+. "
                "Branded queries rank pos ~6; non-brand avg pos ~46. See the 'Goal Validation' tab.")
    ws["A5"].font = CELL_FONT
    toc = [
        ("Goal Validation", "Why nothing ranks: brand vs non-brand, position distribution, referring-domain quality, levers"),
        ("Recommendations", "Prioritized P0/P1/P2 actions"),
        ("Pages (overview)", "All 277 pages — core per-page metrics"),
        ("Crawlability", "Keyword pages that are not indexable"),
        ("Titles", "Title tags outside 15-60 chars"),
        ("Meta Descriptions", "Descriptions outside 70-160 chars / missing"),
        ("Headings", "Pages with H1 != 1"),
        ("Schema Issues", "JSON-LD correctness / usability problems"),
        ("Internal Linking", "Orphans, almost-orphans, poor inbound/outbound (link graph)"),
        ("Redirects", "Broken (active 301 -> 404) and stray (dead) redirect rules"),
        ("Anchor Mismatches", "Internal anchors pointing to a weaker destination (deep, confidence-labeled)"),
        ("Keyword Targeting", "Primary/secondary targeting issues (deep, confidence-labeled)"),
        ("Cannibalization", "Pages competing for the same keyword (deep, confidence-labeled)"),
        ("Keyword Frequency", "Under-optimized primary density (landing -> list -> other)"),
    ]
    ws.append([]); ws.append([])
    r = ws.max_row + 1
    ws.cell(r, 1, "Tab"); ws.cell(r, 2, "Contents")
    for c in ws[r]:
        c.fill = PatternFill("solid", fgColor=NAVY); c.font = HEADER_FONT
    for name, desc in toc:
        ws.append([name, desc])
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 95

    # ---- Goal Validation ----
    ws = wb.create_sheet("Goal Validation")

    def block(ws, title, headers, rows, start):
        ws.cell(start, 1, title).font = SUB_FONT
        hr = start + 1
        for j, h in enumerate(headers, 1):
            c = ws.cell(hr, j, h); c.fill = PatternFill("solid", fgColor=NAVY); c.font = HEADER_FONT
        for i, row in enumerate(rows, hr + 1):
            for j, v in enumerate(row, 1):
                ws.cell(i, j, v).font = CELL_FONT
        return hr + 1 + len(rows) + 2

    ws.cell(1, 1, "Conclusion — why the site isn't ranking").font = TITLE_FONT
    nxt = 3
    b, n = g["brand"], g["non"]
    nxt = block(ws, "Brand vs non-brand (6-week GSC)",
                ["Query type", "Query-rows", "Clicks", "Avg position"],
                [["Branded (piperocket...)", b["rows"], b["clicks"], round(b["pos"], 1)],
                 ["Non-brand (the content program)", n["rows"], n["clicks"], round(n["pos"], 1)]], nxt)
    ti = g["tot_impr"]
    nxt = block(ws, "Impressions by position bucket",
                ["Position bucket", "Impressions", "Share %"],
                [[k, v, round(v / ti * 100, 1)] for k, v in g["buckets"].items()], nxt)
    nxt = block(ws, f"Referring domains by Authority Score (Semrush; AS {SEMRUSH_AS}, {SEMRUSH_RD} total)",
                ["Authority Score", "Referring domains", "Share", "Worth"],
                [[rng, cnt, pct, note] for rng, cnt, pct, note in SEMRUSH_RD_DIST], nxt)
    nxt = block(ws, "What actually moves rankings (priority order)",
                ["#", "Lever", "Why"],
                [[1, "Backlinks / digital PR (RDs ~50 -> 150+)", "The bottleneck. Quality AS40+ editorial links: data studies, founder-led/HARO, write-for-us, partnerships."],
                 [2, "Target within weight class", "Keywords whose page-1 competitors are AS <=25-30 (long-tail, segment-specific, BOFU)."],
                 [3, "Consolidate topical authority", "Dominate ONE cluster first instead of spreading across 240 pages."],
                 [4, "Stop equity leaks", "Fix broken redirects; fix orphaned money pages. Free, but not enough alone."]], nxt)
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 90
    ws.column_dimensions["D"].width = 16

    # ---- Recommendations ----
    sheet(wb, "Recommendations", ["Priority", "Recommendation", "Effort"], [
        ["P0", "Add 'compare' to $isArticle in head-meta.html — 11 compare pages emit no Article/BreadcrumbList schema", "1 template"],
        ["P0", "Set featuredImage on ~106 Article pages using the fallback logo SVG", "Batch/template"],
        ["P0", "Repoint the 27 'early-stage startups' glossary anchors to /list/best-saas-seo-agencies-for-startups/", "1 batch edit"],
        ["P1", "Fix broken redirects (legacy URLs -> 404 slugs)", "Per Redirects tab"],
        ["P1", "Add in-content inbound links to orphan pages (esp. listicles + landing)", "Per Internal Linking tab"],
        ["P1", "Retarget /compare/ pages from zero-demand 'X vs Y' to competitor-review demand", "Per Keyword Targeting"],
        ["P1", "Resolve confirmed cannibalization (merge/canonical/differentiate)", "Per Cannibalization tab"],
        ["P2", "Lift primary density on landing pages then listicles below 3 mentions", "Per Keyword Frequency"],
        ["P2", "Trim long titles / meta descriptions into SERP window; fix what-is-crawling H1", "Cosmetic"],
        ["P-OFF", "Off-page: this is the real lever — build AS40+ referring domains (see Goal Validation)", "Ongoing"],
    ], widths={"A": 10, "B": 95, "C": 22}, wrap=("B",))

    # ---- Pages overview ----
    prows = []
    for p in kw + [p for p in pages if not p["keyword_target"]]:
        gg = cmeta.get(p["url"], {}).get("gsc", {})
        prows.append([p["url"], p["type"], p["keyword_target"], p["primary"], p["funnel"], p["intent"],
                      p["cluster"] or "", p["indexable"], p["title_len"], p["desc_len"], p["h1_count"],
                      p["h2_count"], p["schema_issue_count"] if "schema_issue_count" in p else len(p["schema_issues"]),
                      p["rendered_word_count"], p["primary_freq_exact"], p["primary_freq_family"],
                      p["internal_link_count"], p["external_link_count"], p["inbound_count"], p["outbound_count"],
                      len(p["cannibalization"]), gg.get("top_query", ""), gg.get("top_query_position", ""),
                      gg.get("primary_position", ""), gg.get("aligned", "")])
    sheet(wb, "Pages (overview)",
          ["url", "type", "keyword_target", "primary", "funnel", "intent", "cluster", "indexable",
           "title_len", "desc_len", "h1", "h2", "schema_issues", "rendered_words", "primary_exact",
           "primary_family", "internal", "external", "inbound", "outbound", "cannib_count",
           "gsc_top_query", "gsc_top_pos", "gsc_primary_pos", "gsc_aligned"],
          prows, widths={"A": 42, "D": 26, "G": 18, "V": 26}, wrap=())

    # ---- Crawlability ----
    bad = [[p["url"], p["type"], p["noindex"], p["in_sitemap"], p["canonical_self"], p["is_alias"]]
           for p in kw if not p["indexable"]]
    sheet(wb, "Crawlability", ["url", "type", "noindex", "in_sitemap", "canonical_self", "is_alias"],
          bad or [["All keyword pages indexable", "", "", "", "", ""]], widths={"A": 42})

    # ---- Titles ----
    sheet(wb, "Titles", ["url", "type", "title_len", "title"],
          [[p["url"], p["type"], p["title_len"], p["title"]] for p in pages
           if p["title_len"] > 60 or (p["rendered"] and p["title_len"] < 15)],
          widths={"A": 42, "D": 70}, wrap=("D",))

    # ---- Meta Descriptions ----
    drows = []
    for p in pages:
        if p["desc_len"] > 160 or (p["rendered"] and 0 < p["desc_len"] < 70) or (p["rendered"] and p["desc_len"] == 0 and not p["is_alias"]):
            issue = "too long" if p["desc_len"] > 160 else ("missing" if p["desc_len"] == 0 else "too short")
            drows.append([p["url"], p["type"], p["desc_len"], issue, p["description"]])
    sheet(wb, "Meta Descriptions", ["url", "type", "desc_len", "issue", "description"], drows,
          widths={"A": 42, "E": 70}, wrap=("E",))

    # ---- Headings ----
    sheet(wb, "Headings", ["url", "type", "h1_count", "h1s", "h2_count"],
          [[p["url"], p["type"], p["h1_count"], " | ".join(p.get("h1_texts", [])), p["h2_count"]]
           for p in pages if p["rendered"] and not p["is_alias"] and p["h1_count"] != 1],
          widths={"A": 42, "D": 60}, wrap=("D",))

    # ---- Schema Issues ----
    sheet(wb, "Schema Issues", ["url", "type", "schema_types", "issue_count", "issues"],
          [[p["url"], p["type"], "|".join(p["schema_types"]), len(p["schema_issues"]), "; ".join(p["schema_issues"])]
           for p in pages if p["schema_issues"]],
          widths={"A": 42, "C": 28, "E": 70}, wrap=("E",))

    # ---- Internal Linking ----
    lrows = []
    for p in kw:
        if p["inbound_count"] == 0:
            bucket = "orphan (0 inbound)"
        elif p["inbound_count"] == 1:
            bucket = "almost-orphan (1)"
        elif p["inbound_count"] == 2:
            bucket = "poor inbound (2)"
        elif p["body_word_count"] > 600 and p["outbound_count"] < 2:
            bucket = "poor outbound"
        else:
            continue
        lrows.append([bucket, p["url"], p["type"], p["inbound_count"], p["outbound_count"], p["body_word_count"]])
    lrows.sort(key=lambda r: (r[0], r[3]))
    sheet(wb, "Internal Linking", ["bucket", "url", "type", "inbound", "outbound", "body_words"],
          lrows, widths={"A": 20, "B": 42})

    # ---- Redirects ----
    rrows = [["BROKEN (-> 404)", s, t] for s, t in rh.get("broken", [])] + \
            [["STRAY (dead rule)", s, t] for s, t in rh.get("stray", [])]
    sheet(wb, "Redirects", ["kind", "source", "target"], rrows or [["none", "", ""]],
          widths={"A": 18, "B": 55, "C": 55})

    # ---- Anchor Mismatches (deep) ----
    arows = []
    for url, items in anchors.items():
        for a in items:
            arows.append([url, a.get("confidence", ""), a["anchor"], a["href"], a["better_target"] or "", a["why"]])
    arows.sort(key=lambda r: {"high": 0, "medium": 1, "low": 2}.get(r[1], 3))
    sheet(wb, "Anchor Mismatches", ["page", "confidence", "anchor", "current_href", "better_target", "why"],
          arows, widths={"A": 40, "C": 26, "D": 34, "E": 34, "F": 70}, wrap=("F",))

    # ---- Keyword Targeting (deep) ----
    krows = []
    for url, items in keywords.items():
        for k in items:
            krows.append([url, k.get("confidence", ""), k["problem"], k["detail"], k["suggestion"]])
    krows.sort(key=lambda r: {"high": 0, "medium": 1, "low": 2}.get(r[1], 3))
    sheet(wb, "Keyword Targeting", ["page", "confidence", "problem", "detail", "suggestion"],
          krows, widths={"A": 40, "C": 20, "D": 70, "E": 60}, wrap=("D", "E"))

    # ---- Cannibalization (deep) ----
    crows = []
    for url, items in cannib.items():
        for x in items:
            crows.append([url, x.get("confidence", ""), x.get("verdict", ""), x["severity"],
                          x["recommended_action"], x["competitor"], x["why"]])
    crows.sort(key=lambda r: {"high": 0, "medium": 1, "low": 2}.get(r[1], 3))
    sheet(wb, "Cannibalization", ["page", "confidence", "verdict", "severity", "action", "competitor", "why"],
          crows, widths={"A": 40, "F": 40, "G": 60}, wrap=("G",))

    # ---- Keyword Frequency ----
    frows = []
    for p in [x for x in kw if x["type"] == "landing" and x["primary"] and x["primary_freq_family"] < 3]:
        frows.append(["1-landing", p["url"], p["type"], p["rendered_word_count"], p["primary_freq_exact"], p["primary_freq_family"], p["primary"]])
    for p in [x for x in kw if x["type"] == "list" and x["primary_freq_family"] < 3]:
        frows.append(["2-list", p["url"], p["type"], p["rendered_word_count"], p["primary_freq_exact"], p["primary_freq_family"], p["primary"]])
    for p in [x for x in kw if x["type"] not in ("landing", "list") and x["body_word_count"] > 800 and x["primary_freq_family"] < 2]:
        frows.append(["3-other", p["url"], p["type"], p["rendered_word_count"], p["primary_freq_exact"], p["primary_freq_family"], p["primary"]])
    sheet(wb, "Keyword Frequency", ["priority", "url", "type", "rendered_words", "primary_exact", "primary_family", "primary"],
          frows, widths={"A": 12, "B": 42, "G": 30})

    wb.save(OUT)
    print(f"Wrote {OUT}  ({len(wb.sheetnames)} sheets: {', '.join(wb.sheetnames)})")


if __name__ == "__main__":
    main()
